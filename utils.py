import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import random 

def create_color_palettes(N, seed = 42):
    """
    Create a dictionary with the color palettes for the profiles.
    """
    color_palettes = set() 
    np.random.seed(seed)
    while len(color_palettes) < N:
        color = "#{:02X}{:02X}{:02X}".format(*np.random.randint(128, 255, size=(3,)))
        color_palettes.add(color)
    dict_color = dict(zip(range(1,N+1), color_palettes))
    return dict_color

def hex_to_argb(hex_color):
    """
    Convert a hex color to ARGB. This is need for the excel file. 
    """
    hex_color = hex_color.lstrip("#")
    return "FF" + hex_color

def create_random_data_frame(N, pNperiods, pNpeople):
    data = np.random.randint(1, N + 1, size=(pNpeople, pNperiods))  
    df = pd.DataFrame(data)
    return df

def apply_color(value, color_palettes):
    """
    Apply the color to the cell in the excel file.
    """
    if value == 0:
        return None
    color_ref = color_palettes[value]
    color = hex_to_argb(color_ref)
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    return fill

def access_model_variables(name_variable, index_variables, model): 
    """
    Access to the variables of the model.
    """
    variable = {}
    if index_variables == 1: 
        for var in model.getVars():
            if name_variable in var.VarName:
                i = int(var.VarName.split('_')[1])
                variable[i] = var
    elif index_variables == 2: 
        for var in model.getVars():
            if name_variable in var.VarName:
                i, t = map(int, var.VarName.split('_')[1:])
                variable[i, t] = var
    elif index_variables == 3:
        for var in model.getVars():
            if name_variable in var.VarName:
                i, j, t = map(int, var.VarName.split('_')[1:])
                variable[i, j, t] = var
    return variable

def add_sheet_excel(excel_file, name_sheet, df_data, index = False): 
    """
    Add a sheet to the excel file.
    """
    if name_sheet in openpyxl.load_workbook(excel_file).sheetnames:
        wb = openpyxl.load_workbook(excel_file)
        wb.remove(wb[name_sheet])
        wb.save(excel_file)
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        df_data.to_excel(writer, sheet_name=name_sheet, index = index)

def columns_dimensions(excel_file, wb, sheet, df, width = 10):
    """
    Modify the width of the columns in the excel file.
    """
    for i in range(df.shape[1]+1):
        column_letter = get_column_letter(i+1)
        sheet.column_dimensions[column_letter].width = width
    wb.save(excel_file)

def weighted_random_choice(probabilities):
    """
    Selects a number based on given probabilities.
    """
    numbers = [0, 1, 2]
    return np.random.choice(numbers, p=probabilities)

def modify_list(input_list, j):
    """
    Modify the list with ones. It makes that in the least can be at maximun two ones.
    """
    if len(input_list) <= 1:
        return input_list 
    one_indices = [i for i, x in enumerate(input_list[j:], start=1) if x == 1]
    if not one_indices:
        return input_list 
    keep_index = random.choice(one_indices)
    for i in one_indices:
        if i != keep_index:
            input_list[i] = 0
    return input_list
   
def generate_probabilities(number_preference):
    """
    Generate the probabilities for the availability of the people.
    """
    highest_probability = random.uniform(0.6, 0.8)
    remaining_probability = 1 - highest_probability
    probabilities = [0, 0, 0]
    second_probability = random.uniform(0, remaining_probability)
    third_probability = remaining_probability - second_probability
    if second_probability > third_probability:
        second_highest_probability, third_highest_probability = second_probability, third_probability
    else: 
        second_highest_probability, third_highest_probability = third_probability, second_probability
    if number_preference == 0:
        probabilities[number_preference] = highest_probability
        probabilities[1] = second_highest_probability
        probabilities[2] = third_highest_probability
    elif number_preference == 1:
        probabilities[number_preference] = highest_probability
        probabilities[0] = second_highest_probability
        probabilities[2] = third_highest_probability
    elif number_preference == 2:
        probabilities[number_preference] = highest_probability
        probabilities[1] = second_highest_probability
        probabilities[0] = third_highest_probability
    return probabilities

def sorted_profiles(excel_file):
    """
    Sort the profiles based on the number of people that have the profile.
    """
import pandas as pd

# Sample DataFrame creation (you would load your actual data)
data = {
    'Period 1': ["", "OTS", "OTS", "MDCA", "SAR"],
    'Period 2': ["OTS", "OTS", "", "MDCA", "SAR"],
    'Period 3': ["OTS", "OTS", "GS", "MDCA", "SAR"],
    'Period 4': ["", "OTS", "GS", "MDCA", ""],
    'Period 5': ["", "", "", "", "GS"]
}
index = ["Person 2", "Person 3", "Person 4", "Person 8", "Person 11"]

df = pd.DataFrame(data, index=index)

def find_first_period(row):
    """
    Find the first period that the person is attending the emergency.
    """
    for i, item in enumerate(row):
        if item:  
            return i

def find_last_period(row, pNperiods):
    """
    Find the last period that the person is attending the emergency.
    """
    for i in range(5,0, -1):  
        if row[i]:  
            return i
    return None  

def sorted_profiles(df, pNperiods):
    """
    Sort the profiles based on the periods the person is attending the emergency.
    """
    df['First_Period'] = df.apply(find_first_period, axis=1)
    df['Last_Period'] = df.apply(find_last_period, axis=1, pNperiods = pNperiods)

    df_sorted = df.sort_values(by=['First_Period', 'Last_Period'])
    df_sorted = df_sorted.drop(columns=['First_Period', 'Last_Period'])
    return df_sorted

def to_latex(datafile, sheet): 
    """
    Convert the excel file to latex code.
    """
    df = pd.read_excel(datafile, sheet_name=sheet)
    latex = df.to_latex(index=False)
    latex = df.fillna('')
    print(latex.to_latex(index=False))

