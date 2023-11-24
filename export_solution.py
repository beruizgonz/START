import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

N = 10  # Change N to the desired number of color palettes

def create_color_palettes(N=N):
    color_palettes = set()  # Use a set to ensure uniqueness
    while len(color_palettes) < N:
        # Generate a random color palette
        color = "#{:02X}{:02X}{:02X}".format(*np.random.randint(0, 255, size=(3,)))
        color_palettes.add(color)
    return list(color_palettes)


def hex_to_argb(hex_color):
    hex_color = hex_color.lstrip("#")
    return "FF" + hex_color

def create_random_data_frame(N=N):
    data = np.random.randint(0, N, size=(18, 20))
    df = pd.DataFrame(data)
    return df

def apply_color(value, color_palettes):
    color = hex_to_argb(color_palettes[value])
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    return fill

def export_to_excel(df, color_palettes):
    # Create a Workbook
    workbook = Workbook()
    sheet = workbook.active

    # Apply colors to the entire DataFrame and export to Excel
    for i, row in enumerate(df.values):
        for j, cell_value in enumerate(row):
            color_fill = apply_color(cell_value -1, color_palettes)
            # addt the value to the cell
            if cell_value == 0:
                sheet.cell(row=i + 1, column=j + 1).value = ''
                # sheet.cell(row=i + 1, column=j + 1).fill = None
            else:
                sheet.cell(row=i + 1, column=j + 1).value = cell_value
                sheet.cell(row=i + 1, column=j + 1).fill = color_fill
           

    workbook.save('planning1.xlsx')

if __name__ == "__main__":
    N = 6
    color_palettes = create_color_palettes(6)
    df = create_random_data_frame(6)
    export_to_excel(df, color_palettes)




