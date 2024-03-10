import pandas as pd
import openpyxl
import os
import random as random
import numpy as np

from utils import add_sheet_excel, columns_dimensions, generate_probabilities, weighted_random_choice, modify_list
from opts import parser_args

def create_excel_file(opts):
    wb = openpyxl.Workbook()
    wb.active.title = 'Data'
    sheet = wb['Data']
    sheet['A1'] = 'N periods'
    sheet['B1'] = opts.nPeriods
    sheet['A2'] = 'N health profiles'
    sheet['B2'] = opts.nProfiles
    sheet['A3'] = 'N people'
    sheet['B3'] = opts.nPeople
    sheet['A4'] = 'Number chartered'
    sheet['B4'] = opts.nCharter
    sheet['A5'] = 'Discount'
    sheet['B5'] = opts.Discount
    sheet['A6'] = 'N people discount'
    sheet['B6'] = opts.nDiscount
    sheet['A7'] = 'Max periods'
    sheet['B7'] = opts.maxPeriods
    sheet['A8'] = 'Min periods'
    sheet['B8'] = opts.minPeriods
    sheet.column_dimensions['A'].width = 18
    wb.save(opts.output_path)

def create_prices(opts):
    """
    Create the sheet 'Prices' in the excel file. The prices are genereated with an uniform distribution based on real prices.
    """
    df_prices = pd.DataFrame(index=range(1,opts.nPeriods +1), columns=['Outward', 'Return'])
    random.seed(opts.seed)
    for i in range(opts.nPeriods):
        df_prices.loc[i+1, 'Outward'] = random.randint(100, 1000)
        df_prices.loc[i+1, 'Return'] = random.randint(100, 1000)
    df_prices['Tperiod'] = df_prices.index
    df_prices = df_prices[['Tperiod', 'Outward', 'Return']]  
    add_sheet_excel(opts.output_path, 'Prices', df_prices, False)

def chartered_flights(opts):
    """
    Create the sheet 'Chartered' in the excel file. The prices of the charter flights are fixed.
    """
    df = pd.DataFrame(index=range(1, opts.nPeriods + 1), columns=[f'Chartered {i + 1}' for i in range(opts.nCharter)])
    random.seed(opts.seed)
    df['Chartered 1'] = opts.CostCharter1
    df ['Chartered 2'] = opts.CostCharter2
    min_cap_values = opts.minCapacity
    max_cap_values = opts.maxCapacity
    df.insert(2, '', '') 
    df['Min Cap'] = min_cap_values[:len(max_cap_values)] + [None] * (opts.nPeriods - len(max_cap_values))
    df['Max Cap'] = max_cap_values[:len(max_cap_values)] + [None] * (opts.nPeriods - len(max_cap_values))
    df['Tperiod'] = df.index
    df = df[['Tperiod', 'Chartered 1', 'Chartered 2', '', 'Min Cap', 'Max Cap']]
    add_sheet_excel(opts.output_path, 'Charter', df, False)
    wb = openpyxl.load_workbook(opts.output_path)
    sheet = wb['Charter']
    columns_dimensions(opts.output_path, wb, sheet, df, width = 14)

def demand(opts):
    """
    Create the sheet 'Demand' in the excel file. The demand of the emergency is generated based on the profiles of a EMT2 hospital.
    """
    wb = openpyxl.load_workbook( opts.profile_path)
    sheet = wb['EMT 2']
    df_profiles = pd.DataFrame(sheet.values)
    df = pd.DataFrame(index = range(1,opts.nProfiles + 1), columns= range(1,opts.nPeriods + 1))
    for i in range(opts.nProfiles):
        df.loc[i+1,:] = df_profiles.iloc[i+1, 3]
    df['Profile | Tperiod'] = df.index
    df = df[['Profile | Tperiod'] + list(df.columns[:-1])]
    add_sheet_excel(opts.output_path, 'Demand', df, False)
    wb = openpyxl.load_workbook(opts.output_path)
    sheet = wb['Demand']
    columns_dimensions(opts.output_path, wb, sheet, df, width = 14)

def health_profiles(opts):
    """
    Create the sheet 'HealthProfiles' in the excel file. The health profiles are created based on a compatibility matrix
    that dictates which types of profiles can be performed by a person. That is, there are different profiles, and a person 
    has a main profile that they will primarily perform. Depending on that profile, they may or may not be able to perform
    another.
    """
    wb = openpyxl.load_workbook(opts.profile_path)
    sheet = wb['EMT 2']
    df_profiles = pd.DataFrame(sheet.values)
    abbreviation = df_profiles.iloc[1:, 2]
    dict_profiles = {}
    sheet_probability = wb['Probabilities']
    df_probability = pd.DataFrame(sheet_probability.values)
    df_probability = df_probability.drop(0)
    df_probability = df_probability.drop(0, axis=1)
    df_profiles = df_profiles.drop(0)
    for i in range(1, opts.nProfiles + 1):
        dict_profiles[i] = df_profiles.iloc[i-1, 3]
    df_health_profiles = pd.DataFrame(columns=range(1, opts.nProfiles + 1))
    k = 0
    #for i, row in df_probability.iterrows():
    # read only the opt.nProfiles first rows
    for i, row in df_probability.iloc[:opts.nProfiles].iterrows():
        profile = dict_profiles[i]
        profile = int(profile)
        for j in range(opts.ratio*profile):
            binomial_results = [np.random.binomial(1, float(p)) for p in row[:opts.nProfiles]]
            if i == 1 or i == 2:
                binomial_results = modify_list(binomial_results, i)
            df_health_profiles.loc[k,:] = binomial_results
            k += 1
    df_health_profiles['Person | Profile'] = df_health_profiles.index +1
    abb = pd.DataFrame([abbreviation]) 
    # Get only the opt.nProfiles first rows
    abb = abb.iloc[:opts.nProfiles]
    df_health_profiles = pd.concat([abb,df_health_profiles]).reset_index(drop=True)
    df_health_profiles = df_health_profiles[['Person | Profile'] + list(df_health_profiles.columns[:opts.nProfiles])]
    add_sheet_excel(opts.output_path, 'HealthProfiles', df_health_profiles, False)  
    wb = openpyxl.load_workbook(opts.output_path)
    sheet = wb['HealthProfiles']
    columns_dimensions(opts.output_path, wb, sheet, df_health_profiles, width = 20)

def availability(opts):
    """
    Create the sheet 'Availability' in the excel file. The availability of the people is generated based on
    their number preference.
    """
    df = pd.DataFrame(index=range(1, opts.nPeople + 1), columns=range(1, opts.nPeriods + 1))
    for i in range(opts.nPeople):
        # number_preference = random.randint(0, 2) # Example with a uniform distribution
        # number_preference = random.choices([0, 1, 2], [0.5, 0.3,0.2], k=1)[0] # Example with a non-uniform distribution and less 2s
        number_preference = random.choices([0, 1, 2], [0.4,0.4,0.2], k=1)[0] # Example with a non-uniform distribution and less 2s but a higher probability for 2
        probabilities = generate_probabilities(number_preference)
        #print(f'Person {i+1}: Number Preference = {number_preference}, Probabilities = {probabilities}')
        for j in range(opts.nPeriods):
            df.loc[i+1, j+1] = weighted_random_choice(probabilities)
    add_sheet_excel(opts.output_path, 'Availability', df, True)

def note(opts): 
    """
    Create the sheet 'Note' in the excel file. The note is a paramter of the model in case of tie. 
    """
    df = pd.DataFrame(index=range(1, opts.nPeople + 1), columns= ['Grade'])
    for i in range(opts.nPeople):
        # Create a function of probability. I have three values: 0, 1, 2. For value 0 the probability is 0.1, for value 1 the probability is 0.2 and for value 2 the probability is 0.7.
        numbers = [0, 1, 2]
        probabilities = [0.4, 0.5, 0.1]
        number_select = random.choices(numbers, probabilities, k=1)[0]
        if number_select == 0:
            note = random.uniform(5,6.999999)
        elif number_select == 1:
            note = random.uniform(7,8.999999)
        else:
            note = random.uniform(9, 10)
        note = round(note, 2)
        df.loc[i+1, 'Grade'] = note
    add_sheet_excel(opts.output_path, 'Grades', df, True)

# def weights(opts):
#     """
#     Define the weights for the objective function.
#     """
#     df = pd.DataFrame(index=range(1, 6), columns=['Weight', 'Value ', 'Description'])
#     df.loc[1, 'Weight'] = 'W1'
#     df.loc[1, 'Value '] = opts.w1
#     df.loc[1, 'Description'] = 'Weight for the cost'
#     df.loc[2, 'Weight'] = 'W2'
#     df.loc[2, 'Value '] = opts.w2
#     df.loc[2, 'Description'] = 'Weight for infeasibility'
#     df.loc[3, 'Weight'] = 'W3'
#     df.loc[3, 'Value '] = opts.w3
#     df.loc[3, 'Description'] = 'Weight for second objective'
#     df.loc[5, 'Weight'] = 'ww1'
#     df.loc[5, 'Value '] = opts.ww1
#     df.loc[5, 'Description'] = 'Weight for chartered flights'
#     df.loc[6, 'Weight'] = 'ww2'
#     df.loc[6, 'Value '] = opts.ww2
#     df.loc[6, 'Description'] = 'Weight for regular flights'
#     df.loc[7, 'Weight'] = 'ww3'
#     df.loc[7, 'Value '] = opts.ww3
#     df.loc[7, 'Description'] = 'Weight for number of roles'
#     df.loc[8, 'Weight'] = 'ww4'
#     df.loc[8, 'Value '] = opts.ww4
#     df.loc[8, 'Description'] = 'Weight for availability'
#     add_sheet_excel(opts.output_path, 'Weights', df, False)
#     wb = openpyxl.load_workbook(opts.output_path)
#     sheet = wb['Weights']
#     sheet.delete_rows(1)
#     columns_dimensions(opts.output_path, wb, sheet, df, width = 30)


if __name__ == '__main__':
    opt = parser_args()
    create_excel_file(opt)
    demand(opt)
    create_prices(opt)
    chartered_flights(opt)
    health_profiles(opt)
    availability(opt)
    note(opt)



