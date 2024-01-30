import os
import gurobipy as gp
from gurobipy import GRB
import pandas as pd
import numpy as np
import openpyxl
from STARTmodel import START

from utils import create_color_palettes, apply_color, access_model_variables, add_sheet_excel, columns_dimensions
#  Read the data from the excel file
datafile='Data2Inf.xlsx'

shdata = pd.read_excel(datafile, sheet_name='Data', header=None)
pNperiods = shdata.loc[0,1]
pNhealthp = shdata.loc[1,1]
pNpeople = shdata.loc[2,1]
pNcharter = shdata.loc[3,1]
pDiscount = shdata.loc[4,1]
pKpeople = shdata.loc[5,1]
pMaxperiods = shdata.loc[6,1]
pMinperiods = shdata.loc[7,1]

shprices = pd.read_excel(datafile, sheet_name='Prices')
pCostOut = shprices.loc[:,'Outward'].to_numpy()
pCostRet = shprices.loc[:,'Return'].to_numpy()

shcharter = pd.read_excel(datafile, sheet_name='Charter')
pCostChar = shcharter.loc[0:pNperiods-1,'Chartered 1':'Chartered 2'].to_numpy()
pMinCapCh = shcharter.loc[0:pNcharter-1,'Min Cap'].to_numpy()
pMaxCapCh = shcharter.loc[0:pNcharter-1,'Max Cap'].to_numpy()

shhealthprofiles = pd.read_excel(datafile, sheet_name='HealthProfiles', header=None)
pNameprofiles = shhealthprofiles.loc[0,1:].to_numpy()
pNameabbprofiles = shhealthprofiles.loc[1,1:].to_numpy()
pProfiles = shhealthprofiles.loc[2:,1:].to_numpy()

def model_dimensions(model, excel):
    """
    Create the sheet 'Model dimensions' in the excel file.
    The sheet contains the number of variables, constraints and the type of variables.
    """
    df = pd.DataFrame(index = ['total', 'continuous', 'binary', 'integer', 'constrains'], columns=['Type', 'Number'])
    df['Type'] = df.index
    df['Number'] = [    
        model.NumVars,
        model.NumVars-model.NumIntVars,
        model.NumBinVars,
        model.NumIntVars-model.NumBinVars, 
        model.NumConstrs
    ]
    add_sheet_excel(excel, 'Model dimensions', df, index = True)
    wb = openpyxl.load_workbook(excel)
    sheet = wb['Model dimensions']
    columns_dimensions(excel, wb, sheet, df, 16)

def model_performance(model,excel): 
    info = ['Objective value', 'Runtime', 'MIPGap', 'Status']
    df = pd.DataFrame(index = info, columns=['Value'])
    df['Value'] = [
        model.objVal,
        model.Runtime,
        model.MIPGap,
        model.Status
    ]
    add_sheet_excel(excel, 'Model performance', df, index = True)

def chartered(model, excel):
    """
    Create the sheet 'Charter' in the excel file. The sheet contains the number of people that are chartered and the cost.
    """ 
    vGamma = access_model_variables('vGamma', 2, model)
    charter_data = []
    cost_data = []
    vGamma_array = np.array([[vGamma[t, l].X for l in range(pNcharter)] for t in range(pNperiods)])
    for l in range(pNcharter):
        nl = np.sum(vGamma_array[:, l])
        costl = np.sum(pCostChar[:, l] * vGamma_array[:, l])
        charter_data.append(nl)
        cost_data.append(costl)
    df = pd.DataFrame(index = ['Chartered 1', 'Chartered 2'], columns=['Number', 'Cost'])
    df['Number'] = charter_data
    df['Cost'] = cost_data
    add_sheet_excel(excel, 'Charter', df, index = True)

def regular_and_cost(excel, model): 
    """
    Create the sheet 'Regular' in the excel file. The sheet contains the total cost of the outward and return flights and the 
    sum of both.
    """
    costoutflights = 0
    costretflights = 0
    vXstand = access_model_variables('Xstand', 1, model)
    vXdisc = access_model_variables('Xdisc', 1, model)
    vYstand = access_model_variables('Ystand', 1, model)
    vYdisc = access_model_variables('Ydisc', 1, model)
    for t in range(pNperiods - 1):
        costoutflights += pCostOut[t] * vXstand[t].X + (pCostOut[t] - pDiscount) * vXdisc[t].X
        costretflights += pCostRet[t + 1] * vYstand[t + 1].X + (pCostRet[t + 1] - pDiscount) * vYdisc[t + 1].X
    totcostreg = costoutflights + costretflights
    df = pd.DataFrame(index = ['Outward', 'Return','Total'], columns=['Value'])
    df['Value'] = [costoutflights, costretflights, totcostreg]
    add_sheet_excel(excel, 'Regular', df, index = True)

def out_and_return(excel, model):
    """
    Create the sheet 'Outwards and return' in the excel file. The sheet contains the number of people that are
    outwards and return depending on the type of flight.
    """
    vXstand = access_model_variables('Xstand', 1, model)
    vXdisc = access_model_variables('Xdisc', 1, model)
    vYstand = access_model_variables('Ystand', 1, model)
    vYdisc = access_model_variables('Ydisc', 1, model)
    vZout = access_model_variables('Zout', 1, model)
    vZret = access_model_variables('Zret', 1, model) 

    npeopstand = sum(vXstand[t].X for t in range(pNperiods-1))
    npeopdisc = sum(vXdisc[t].X for t in range(pNperiods-1))
    npeopchar = sum(vZout[t].X for t in range(pNperiods-1))

    npeopstandr = sum(vYstand[t].X for t in range(1, pNperiods))
    npeopdiscr = sum(vYdisc[t].X for t in range(1, pNperiods))
    npeopcharr = sum(vZret[t].X for t in range(1, pNperiods))

    df = pd.DataFrame(index = ['Standar', 'Discounted', 'Chartered'], columns=['Number of people outwards', 'Number of people return'])
    df['Number of people outwards'] = [npeopstand, npeopdisc, npeopchar]
    df['Number of people return'] = [npeopstandr, npeopdiscr, npeopcharr]
    add_sheet_excel(excel, 'Outwards and return', df, index = True)
 
def fligths_plan(excel, model): 
    """
    Create the sheet 'Flights plan' in the excel file. The sheet contains the period of departure and return of each person.
    """
    vAlphaout = access_model_variables('Alphaout', 2, model)
    vAlpharet = access_model_variables('Alpharet', 2, model)
    dict = {}
    perdep = 0
    perret = 0
    for i in range(0, pNpeople):
        for t in range(0, pNperiods):
            if (vAlphaout[i,t].X == 1):
                perdep = t
            if (vAlpharet[i,t].X == 1):
                perret = t
        dict[f'Person {i}'] = [perdep+1, perret+1]
    df_people = pd.DataFrame.from_dict(dict, orient='index', columns=['Departure', 'Return'])
    df_people['Person'] =  range(1, len(df_people) + 1)
    df_people = df_people[['Person', 'Departure', 'Return']]  
    add_sheet_excel(excel, 'Flights plan', df_people, index = False)
    wb = openpyxl.load_workbook(excel)
    sheet = wb['Flights plan']
    columns_dimensions(excel, wb, sheet, df_people, 25)

def health_profile_complementary(excel_file ,df, dict_references, color_palettes):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.create_sheet('Profile plan')
    for i in range(1, df.shape[1]):
        sheet.cell(row=1, column=i + 1).value = 'Period ' + str(i)
    for i in range(1, df.shape[0] + 1):
        sheet.cell(row=i + 1, column=1).value = 'Person ' + str(i)
    for i, row in enumerate(df.values):
        for j, cell_value in enumerate(row):
            color_fill = apply_color(cell_value, color_palettes)
            if cell_value == 0:
                sheet.cell(row=i + 2, column=j + 2).value = ''
            else:
                sheet.cell(row=i + 2, column=j + 2).value = dict_references[cell_value] 
                sheet.cell(row=i + 2, column=j + 2).fill = color_fill
    for i, key in enumerate(dict_references):
        sheet.cell(row=i + 5, column=df.shape[1] + 3).value = dict_references[key]
        color_fill = apply_color(key, color_palettes)
        sheet.cell(row=i + 5, column=df.shape[1] + 2).fill = color_fill
    workbook.save(excel_file)

def health_profiles(model, excel_file):
    """
    Create the sheet 'Health profiles' in the excel file. The sheet contains the health profile of each person in each period.
    """
    results = np.zeros((pNpeople, pNperiods), dtype=int)
    vBeta = access_model_variables('vBeta', 3, model)
    for i in range(0,pNpeople):
        for t in range(0,pNperiods):
            results[i,t] = 0
            for j in range (0,pNhealthp):
                if vBeta[i,j,t].X == 1:
                    results[i,t] = j+1
    dict_health_profiles = dict(zip(range(1, pNhealthp + 1), pNameabbprofiles))
    df = pd.DataFrame(results)
    df = df.loc[~(df==0).all(axis=1)]
    color_palettes = create_color_palettes(pNhealthp)
    health_profile_complementary(excel_file, df,dict_health_profiles, color_palettes)

def necessary_people(excel_file, model): 
    """
    Create the sheet 'Necessary people' in the excel file. The sheet contains the number of complementary people that are necessary to 
    attend the emergency in each period.
    """
    vUmas = access_model_variables('vUmas', 2, model)
    df = pd.DataFrame(index = range(0, pNhealthp), columns=range(0, pNperiods))
    dict_health_profiles = dict(zip(range(1, pNhealthp + 1), pNameabbprofiles))
    df_values = np.array([[vUmas[j, t].X for t in range(pNperiods)] for j in range(pNhealthp)])
    df = pd.DataFrame(df_values)
    df.index = dict_health_profiles.values()
    for i in range(0, pNperiods):
        df.rename(columns={i: 'Period ' + str(i+1)}, inplace=True)
    add_sheet_excel(excel_file, 'Necessary people', df, index = True)

if __name__ == '__main__': 
    model = START(datafile)
    model.create_model()
    model.create_variables()
    model.create_constrains()
    model.create_objective_function()
    solution = model.solve()

    filename = f'Res_{datafile}'
    excel_file_path = os.path.join(os.getcwd(), filename)
    wb = openpyxl.Workbook()
    wb.active.title = 'Model dimensions'
    wb.save(excel_file_path)
    model_performance(solution, excel_file_path)
    model_dimensions(solution, excel_file_path)
    chartered(solution, excel_file_path)
    regular_and_cost(excel_file_path, solution)
    out_and_return(excel_file_path, solution)
    fligths_plan(excel_file_path, solution)
    health_profiles(solution, excel_file_path)   
    necessary_people(excel_file_path, solution)
